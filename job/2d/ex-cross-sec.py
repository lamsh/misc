#!/usr/bin/env python3
# coding: utf-8
# (File name: ex-cross-sec.py)
# Author: SENOO, Ken
# (Last update: 2015-01-16T17:51+09:00)
# License: MIT

"""
計算結果から横断ごとに河床高を取得

* 側線の座標 cross_sec.csv
* res0039.ini.dat.plt
後処理をした、ブロック形式のpltファイルをひとまず扱う。

ファイルを読み込む

"""

"""BLOCK形式のtecplotファイルの構造
以下のように、位置でなく変数ごとに列挙する。
X 1 2 3 ... J
  ...I行だけ続く
Y
Z
"""

"""近い値を求める方法
側線の始点･終点から直線を求める。
直線とメッシュの交点を出す
直線をある間隔で分割する。
その分割点ごとに、周りの点を探して平均する

y=ax+b
"""

import sys
import shapely.geometry as geo
import openpyxl



## 側線（kp: kilopost）の始点･終点を取得
FR="./cross_sec.csv"
with open(FR, "U", encoding="utf-8") as fr:
    fr = fr.read().splitlines()
kp_begin = [list(map(float, i.split(",")[1:3])) for i in fr[1:]]
kp_end = [list(map(float, i.split(",")[4:])) for i in fr[1:]]
# kp = zip(kp_begin, kp_end)

kp = []
SHIFT = 500
## 側線を延長
for begin, end in list(zip(kp_begin, kp_end)):
    div = (end[1] - begin[1]) / (end[0] - begin[0])
    kp.append([[begin[0]-SHIFT, begin[1]-div*SHIFT], [end[0]+SHIFT, end[1]+div*SHIFT]])



## 計算結果から値を取得
FR="./res0039.ini.dat.plt"
with open(FR, "U", encoding="utf-8") as fr:
    fr = fr.read().splitlines()

## get variable names
var_names = fr[1].split()[2:]
var_names[:] = [i.strip('",') for i in var_names]

## get max column and row
max_col, max_row = list(map(int, fr[2].split()[2::2]))
var_array={}
for var_i, var in enumerate(var_names):
    if var_i <= 1: # 点の変数
        var_array[var] = [list(map(float, i.split())) for i in fr[4 + var_i * max_row:4 + (var_i + 1) * max_row]]
    else: # セルの変数
        var_array[var] = [list(map(float, i.split())) for i in fr[2+4 + var_i * (max_row - 1) :2+4 + (var_i + 1) * (max_row -1 )]]

## xyの外周を取得
out_xy_right = [[point_x[0], point_y[0]] for point_x, point_y in zip(var_array["X"], var_array["Y"])]
out_xy_left = [[point_x[-1], point_y[-1]] for point_x, point_y in zip(var_array["X"], var_array["Y"])]

## 判定
## 右岸と左岸のラインをリストで用意。そのラインと側線の交差を判定すればよい。
## ただし、側線がメッシュの内側のことがあるので、事前に延長しておく。

## 側線とメッシュ両端の交点を抽出
cross_kp_mesh = []
for kpline in kp:
    gkpline = geo.LineString(kpline)
    gout_xy_right = geo.LineString(out_xy_right)
    gout_xy_left = geo.LineString(out_xy_left)
    cross_right = gkpline.intersection(gout_xy_right)
    cross_left = gkpline.intersection(gout_xy_left)

    ## 交差していなければ無視
    if not cross_right.is_empty and not cross_left.is_empty:
        cross_kp_mesh.append(list(cross_left.coords) + list(cross_right.coords))

# a = geo.LineString(cross_kp_mesh[-1])

## 対岸2点を内挿して値を抽出する座標を用意
cross_distance = []
cross_coord = []
for begin, end in cross_kp_mesh:
    distance = ((end[0]-begin[0])**2 + (end[1]-begin[1])**2)**0.5
    # div = (end[1] - begin[1]) / (end[0] - begin[0])
    div_x = (end[0] - begin[0]) / (max_col-1)
    div_y = (end[1] - begin[1]) / (max_col-1)

    ## 側線での横断距離
    cross_distance.append([i * distance/(max_col-1) for i in range(max_col)])
    ## 値を検索するxy座標
    cross_coord.append([[begin[0] + div_x * col, begin[1] + div_y * col] for col in range(max_col)])
    # ここのrangeをmax_col-1でいいのかな？


## 列方向に検索を掛ける
## セルに含まれていたらそのセルの値を取ることにしよう。
var_x = var_array["X"]
var_y = var_array["Y"]

## 横断と近い行番号を取得
min_row = []
for begin, end in cross_kp_mesh:
    ## 左岸と右岸で小さい行数を探す
    min_num=10000
    for idx, pt in enumerate(out_xy_left):
        error = abs(pt[0] - begin[0]) + abs(pt[1] - begin[1])
        if min_num > error:
            min_num = error
            row_left = idx

    min_num=10000
    for idx, pt in enumerate(out_xy_right):
        error = abs(pt[0] - end[0]) + abs(pt[1] - end[1])
        if min_num > error:
            min_num = error
            row_right = idx

    min_row.append(row_left if row_left < row_right else row_right )



## x座標とy座標の差の絶対値の和が最小のものが最も近いとする。
## 並びの都合上、列の方向と側線の点の向きを同じにする。
"""
1. 側線ごとに
2. 点ごとに
3. セルに含まれるか調査
* 端の列から始めて、見つかれば、次の点を調査。
違う。行から始めて、列を動かす。
次の列で調査。
"""

## 見つかった行･列
cell_row = [[] for i in range(len(cross_coord))]
cell_col = [[] for i in range(len(cross_coord))]

def find_contain():
    for row in range(start_row, max_row-1):
    # for row in range(max_row-1):
        for col in range(max_col-1):
            cell = geo.Polygon([
                (var_x[row][col]    , var_y[row][col]),
                (var_x[row+1][col]  , var_y[row+1][col]),
                (var_x[row+1][col+1], var_y[row+1][col+1]),
                (var_x[row][col+1]  , var_y[row][col+1]),
             ])
            is_contain = cell.contains(pt)
            if is_contain:
                print(cross_idx, row, col)
                cell_row[cross_idx].append(row)
                cell_col[cross_idx].append(col)
                return

## cross_coord[6][-1]
## xy[0] = 1333.73
for cross_idx, xy in enumerate(cross_coord):
    xy=xy[::-1]
    for point in xy:
        pt = geo.Point(point)

        start_row = min_row[cross_idx] -1
        start_row = start_row -1 if start_row == max_row -1 else start_row
        find_contain()

        ## 時間がかかるので関数にしてやってみる。あまり改善されない。
        ## TODO: 速度向上


## 該当する行列がわかったのでその値で抽出する。
## 抽出する値：Z0, Z

# for cross_idx in range(len(cross_coord)):
cross_Z0 = [[] for i in range(len(cross_coord))]
cross_Z  = [[] for i in range(len(cross_coord))]
for cross_idx, (cross_row, cross_col) in enumerate(zip(cell_row,cell_col)):
    for row, col in zip(cross_row, cross_col):
        cross_Z0[cross_idx].append(var_array["Z0"][row][col])
        cross_Z[cross_idx].append(var_array["Z"][row][col])


FW="./out.xlsx"

wb = openpyxl.Workbook()

for cross_idx, (distance, z0, z) in enumerate(zip(cross_distance, cross_Z0, cross_Z)):
    ws = wb.create_sheet(cross_idx)
    ws.title = str(cross_idx)
    ws.cell(row=1, column=1).value = u"横断距離[m]"
    ws.cell(row=1, column=2).value = u"初期河床高[T.P.m]"
    ws.cell(row=1, column=3).value = u"河床高[T.P.m]"
    for row in range(len(z)):
        ws.cell(row=2+row, column=1).value = distance[row]
        ws.cell(row=2+row, column=2).value = z0[row]
        ws.cell(row=2+row, column=3).value = z[row]

wb.save(FW)
